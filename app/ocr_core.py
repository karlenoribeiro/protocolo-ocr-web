# ocr_core.py - ATUALIZADO PARA INCLUIR SETOR/PROFISSIONAL/NOME
import re
import cv2
import numpy as np
import pandas as pd
import pytesseract
from unidecode import unidecode
from datetime import datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from typing import Optional, Tuple, List
from openpyxl import load_workbook
import os

LOCAL_TZ = ZoneInfo("America/Belem")
EXTS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp", ".jfif"}

def now_belem_str():
    return datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S")

def _langs():
    try:
        langs = set(pytesseract.get_languages(config=""))
        return "por+eng" if "por" in langs else "eng"
    except Exception:
        return "por+eng"

OCR_LANGS = _langs()

def to_gray(img):
    if len(img.shape) == 3:
        return cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    return img

def rotate_bound(image, angle):
    (h, w) = image.shape[:2]
    center = (w // 2, h // 2)
    M = cv2.getRotationMatrix2D(center, -angle, 1.0)
    cos = np.abs(M[0, 0]); sin = np.abs(M[1, 0])
    nW = int((h * sin) + (w * cos))
    nH = int((h * cos) + (w * sin))
    M[0, 2] += (nW / 2) - center[0]
    M[1, 2] += (nH / 2) - center[1]
    return cv2.warpAffine(image, M, (nW, nH), flags=cv2.INTER_LINEAR)

def deskew_osd(img):
    try:
        osd = pytesseract.image_to_osd(img)
        angle_m = re.search(r"Rotate: (\d+)", osd)
        if angle_m:
            angle = float(angle_m.group(1)) % 360
            if angle > 1:
                return rotate_bound(img, angle)
    except Exception:
        pass
    return img

def enhance_for_ocr(gray):
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    g = clahe.apply(gray)
    g = cv2.GaussianBlur(g, (3,3), 0)
    _, th = cv2.threshold(g, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    return th

def tesseract_config(digits_only=False):
    base = f"--oem 3 --psm 6 -l {OCR_LANGS}"
    if digits_only:
        base += " -c tessedit_char_whitelist=0123456789-/"
    return base

def ocr_text(img, config):
    return pytesseract.image_to_string(img, config=config)

def ocr_data(img, config):
    return pytesseract.image_to_data(img, output_type=pytesseract.Output.DATAFRAME, config=config)

def find_protocolo_roi(img) -> Optional[Tuple[int,int,int,int]]:
    df = ocr_data(img, tesseract_config(digits_only=False))
    try:
        df = df.dropna(subset=['text'])
    except Exception:
        return None
    if df is None or len(df)==0:
        return None
    df['norm'] = df['text'].astype(str).apply(lambda s: unidecode(s).upper())
    mask = df['norm'].str.contains(r'^PROTOC', regex=True)
    cand = df[mask]
    if cand.empty:
        return None
    cand = cand.sort_values(['top','left']).iloc[0]
    x, y, w, h = int(cand['left']), int(cand['top']), int(cand['width']), int(cand['height'])
    H, W = img.shape[:2]
    pad_y = int(h*2.5)
    pad_x_left = int(w*0.5)
    x1 = max(0, x - pad_x_left)
    y1 = max(0, y - pad_y//2)
    x2 = W
    y2 = min(H, y + pad_y)
    return (x1, y1, x2, y2)

PROTO_REGEXES = [
    re.compile(r'PROTOCOLO[:\s-]*([0-9]{5}\s*[-–]?\s*[0-9]{6}\s*/\s*[0-9]{4}\s*[-–]?\s*[0-9]{2})', re.IGNORECASE),
    re.compile(r'([0-9]{5}\s*[-–]?\s*[0-9]{6}\s*/\s*[0-9]{4}\s*[-–]?\s*[0-9]{2})')
]

def limpar_para_17_digitos(seq: str) -> Optional[str]:
    if not seq:
        return None
    digits = re.sub(r'\D', '', seq)
    return digits if len(digits) == 17 else None

def extrair_protocolo_de_texto(texto: str):
    if not texto:
        return (None, None)
    for rgx in PROTO_REGEXES:
        m = rgx.search(texto)
        if m:
            bruto = m.group(1)
            apenas_numeros = limpar_para_17_digitos(bruto)
            if apenas_numeros:
                fmt = f"{apenas_numeros[0:5]}-{apenas_numeros[5:11]}/{apenas_numeros[11:15]}-{apenas_numeros[15:17]}"
                return (fmt, apenas_numeros)
            return (bruto, None)
    return (None, None)

def tentar_ocr_em_varias_formas(img_color):
    gray1 = to_gray(deskew_osd(img_color))
    th = enhance_for_ocr(gray1)
    inv = cv2.bitwise_not(th)
    morph = cv2.morphologyEx(th, cv2.MORPH_CLOSE, np.ones((3,3), np.uint8), iterations=1)
    cfg = tesseract_config(digits_only=True)
    textos = []
    for candidate in [gray1, th, inv, morph]:
        try:
            textos.append(ocr_text(candidate, cfg))
        except Exception:
            pass
    return "\n".join([t for t in textos if t])

def processar_imagem_bytes(filename: str, content: bytes):
    arr = np.frombuffer(content, dtype=np.uint8)
    img = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if img is None:
        return {
            "arquivo": filename,
            "protocolo_ocr": None,
            "protocolo_17_digitos": None,
            "data_extracao": now_belem_str(),
            "ok": False
        }

    roi = find_protocolo_roi(img)
    textos = ""
    if roi:
        x1,y1,x2,y2 = roi
        recorte = img[y1:y2, x1:x2]
        textos = tentar_ocr_em_varias_formas(recorte)

    if not textos.strip():
        textos = tentar_ocr_em_varias_formas(img)

    protocolo_fmt, protocolo_17 = extrair_protocolo_de_texto(textos)

    if not protocolo_17:
        try:
            texto_extra = ocr_text(img, tesseract_config(digits_only=False))
            protocolo_fmt, protocolo_17 = extrair_protocolo_de_texto(texto_extra)
        except Exception:
            pass

    ok = bool(protocolo_17)
    return {
        "arquivo": filename,
        "protocolo_ocr": protocolo_fmt,
        "protocolo_17_digitos": protocolo_17,
        "data_extracao": now_belem_str(),
        "ok": ok
    }

EXCEL_PATH = Path(__file__).parent / "data" / "protocolos_extraidos.xlsx"
EXCEL_PATH.parent.mkdir(parents=True, exist_ok=True)

def ler_protocolos_existentes() -> Tuple[set, pd.DataFrame]:
    df_existente = pd.DataFrame(columns=[
        "arquivo","protocolo_ocr","protocolo_17_digitos","data_extracao",
        "setor","profissional","nome_completo"
    ])
    if EXCEL_PATH.exists():
        try:
            df_existente = pd.read_excel(EXCEL_PATH, dtype={"protocolo_17_digitos": str})
            # Garante colunas novas mesmo em planilhas antigas
            for col in ["setor","profissional","nome_completo"]:
                if col not in df_existente.columns:
                    df_existente[col] = ""
            df_existente["protocolo_17_digitos"] = (
                df_existente["protocolo_17_digitos"].astype(str).str.strip().replace('', np.nan)
            )
        except Exception:
            pass

    validos = set(
        df_existente[df_existente['protocolo_17_digitos'].str.len() == 17]
        ['protocolo_17_digitos'].dropna().tolist()
    )
    return validos, df_existente

def salvar_excel_texto(df_total: pd.DataFrame):
    if "protocolo_17_digitos" in df_total.columns:
        df_total["protocolo_17_digitos"] = df_total["protocolo_17_digitos"].fillna("").astype(str)
    else:
        df_total["protocolo_17_digitos"] = ""

    # Garante a presença das novas colunas (ordem amigável)
    colunas_final = [
        "arquivo","protocolo_ocr","protocolo_17_digitos","data_extracao",
        "setor","profissional","nome_completo"
    ]
    for col in colunas_final:
        if col not in df_total.columns:
            df_total[col] = ""
    df_total = df_total[colunas_final]

    temp_path = EXCEL_PATH.with_suffix(".tmp")
    try:
        with pd.ExcelWriter(temp_path, engine="openpyxl", mode="w") as writer:
            df_total.to_excel(writer, index=False, sheet_name="Sheet1")
            try:
                ws = writer.book.active
                col_idx = None
                for idx, cell in enumerate(ws[1], start=1):
                    if str(cell.value) == "protocolo_17_digitos":
                        col_idx = idx
                        break
                if col_idx is not None:
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.number_format = "@"
            except Exception:
                pass
        os.replace(temp_path, EXCEL_PATH)
    except Exception as e:
        if temp_path.exists():
            try: temp_path.unlink()
            except Exception: pass
        raise e

def consolidar_resultados(novos: List[dict], df_existente: pd.DataFrame) -> pd.DataFrame:
    df_novos = pd.DataFrame(novos)

    # Colunas padrão + novas
    colunas_padrao = [
        "arquivo","protocolo_ocr","protocolo_17_digitos","data_extracao",
        "setor","profissional","nome_completo"
    ]
    for col in colunas_padrao:
        if col not in df_novos.columns:
            df_novos[col] = None

    df_existente_limpo = df_existente.drop(
        columns=[c for c in df_existente.columns if c not in colunas_padrao],
        errors='ignore'
    )
    df_total = pd.concat(
        [df_existente_limpo, df_novos.drop(columns=['ok','duplicado','indice_processamento'], errors='ignore')],
        ignore_index=True
    )

    if "protocolo_17_digitos" in df_total.columns:
        df_total["protocolo_17_digitos"] = df_total["protocolo_17_digitos"].astype(str).str.strip().replace('', np.nan)

    df_total = df_total[df_total['protocolo_17_digitos'].str.len() == 17].copy()
    df_total = df_total.drop_duplicates(subset=["protocolo_17_digitos"], keep="first")

    salvar_excel_texto(df_total)
    return df_total
