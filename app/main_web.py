# ============================================================
# main_web.py — API FastAPI para OCR de protocolos (17 dígitos)
# - Deduplicação: no lote e contra o Excel existente
# - Gravação como TEXTO no Excel (evita notação científica)
# - Endpoint de download do Excel
# ============================================================

from __future__ import annotations

import time
from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, HTMLResponse
from openpyxl import load_workbook  # opcional (mantido se desejar evoluções)
from datetime import datetime
from zoneinfo import ZoneInfo

# Importa do seu módulo de OCR
from .ocr_core import processar_imagem_bytes, EXCEL_PATH


# -------------------------
# Configurações gerais
# -------------------------
APP_TITLE = "Protocolo OCR Web"
APP_VERSION = "1.2"
LOCAL_TZ = ZoneInfo("America/Belem")

COLS = ["arquivo", "protocolo_ocr", "protocolo_17_digitos", "data_extracao"]


def agora_belem_str() -> str:
    return datetime.now(LOCAL_TZ).strftime("%Y-%m-%d %H:%M:%S")


# -------------------------
# Normalização e salvamento
# -------------------------
def _normaliza_proto_17(s: Any) -> str:
    """Mantém apenas dígitos e aceita somente strings com exatamente 17 dígitos."""
    if pd.isna(s):
        return ""
    s = "".join(ch for ch in str(s) if ch.isdigit())
    return s if len(s) == 17 else ""


def salvar_excel_sem_duplicar(caminho_xlsx: str | Path, df_lote: pd.DataFrame) -> Dict[str, Any]:
    """
    Regras:
    - Garante colunas padrão e que 'protocolo_17_digitos' esteja normalizado (apenas 17 dígitos, texto).
    - Remove duplicados DENTRO do lote por 'protocolo_17_digitos'.
    - Compara com o Excel (se existir) e só ANEXA protocolos realmente novos.
    - Grava coluna como TEXTO no Excel (number_format "@").

    Retorna estatísticas e listas úteis para o front.
    """
    caminho_xlsx = Path(caminho_xlsx)
    caminho_xlsx.parent.mkdir(parents=True, exist_ok=True)

    # --- Normaliza e garante colunas ---
    df = df_lote.copy()
    for c in COLS:
        if c not in df.columns:
            df[c] = ""
    df = df[COLS].astype(str)

    # Normaliza protocolo 17 dígitos
    df["protocolo_17_digitos"] = df["protocolo_17_digitos"].map(_normaliza_proto_17)
    # Remove linhas inválidas (sem 17 dígitos)
    df = df[df["protocolo_17_digitos"] != ""].reset_index(drop=True)

    total_lote = len(df)

    # --- Dedup DENTRO do lote ---
    df_lote_unico = df.drop_duplicates(subset=["protocolo_17_digitos"], keep="first").reset_index(drop=True)
    duplicados_no_lote = total_lote - len(df_lote_unico)

    # Guarda lista dos duplicados do lote (para UI)
    seen = set()
    lista_duplicados_lote: List[str] = []
    for proto in df["protocolo_17_digitos"].tolist():
        if proto in seen and proto not in lista_duplicados_lote:
            lista_duplicados_lote.append(proto)
        seen.add(proto)

    # --- Lê Excel existente (se houver) ---
    if caminho_xlsx.exists():
        df_exist = pd.read_excel(caminho_xlsx, dtype=str, engine="openpyxl")
        for c in COLS:
            if c not in df_exist.columns:
                df_exist[c] = ""
        df_exist = df_exist[COLS].astype(str)
        df_exist["protocolo_17_digitos"] = df_exist["protocolo_17_digitos"].map(_normaliza_proto_17)
        df_exist = df_exist[df_exist["protocolo_17_digitos"] != ""].reset_index(drop=True)
    else:
        df_exist = pd.DataFrame(columns=COLS)

    set_exist = set(df_exist["protocolo_17_digitos"].unique().tolist())
    set_lote_unico = set(df_lote_unico["protocolo_17_digitos"].unique().tolist())

    # Protocolos que JÁ existiam no Excel
    lista_ja_salvos = sorted(list(set_lote_unico & set_exist))
    # Protocolos realmente NOVOS (não existiam no Excel)
    mask_novos = ~df_lote_unico["protocolo_17_digitos"].isin(set_exist)
    df_novos = df_lote_unico[mask_novos].reset_index(drop=True)
    lista_inseridos = df_novos["protocolo_17_digitos"].unique().tolist()

    qtd_ja_salvos = len(lista_ja_salvos)
    inseridos = len(lista_inseridos)

    # --- Concatena e grava ---
    df_final = pd.concat([df_exist, df_novos], ignore_index=True)

    with pd.ExcelWriter(caminho_xlsx, engine="openpyxl", mode="w") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.book["Sheet1"]
        # Força a coluna como TEXTO
        col_idx = list(df_final.columns).index("protocolo_17_digitos") + 1
        for r in range(2, ws.max_row + 1):  # ignora cabeçalho
            ws.cell(row=r, column=col_idx).number_format = "@"

    return {
        "total_lote": total_lote,
        "duplicados_no_lote": duplicados_no_lote,
        "lista_duplicados_lote": lista_duplicados_lote,
        "ja_salvos": qtd_ja_salvos,
        "lista_ja_salvos": lista_ja_salvos,
        "inseridos": inseridos,
        "lista_inseridos": lista_inseridos,
        "total_final": len(df_final),
    }


# -------------------------
# App FastAPI
# -------------------------
app = FastAPI(title=APP_TITLE, version=APP_VERSION)

# CORS — ajuste conforme necessidade (se não usa cookies/autenticação, deixe allow_credentials=False)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],          # em produção, prefira domínios específicos
    allow_credentials=False,      # True exige allow_origins explícito (não "*")
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/", response_class=HTMLResponse)
def root() -> str:
    return "<h1>API Protocolo OCR</h1><p>Use POST /extract</p>"


@app.post("/extract")
async def extract(files: List[UploadFile] = File(...)) -> JSONResponse:
    """
    Recebe imagens, processa OCR e grava Excel sem duplicar protocolos.
    Responde com resumo do lote e caminhos úteis.
    """
    t0 = time.time()
    resultados: List[Dict[str, Any]] = []
    encontrados = 0

    # Processa cada arquivo enviado
    for i, f in enumerate(files, start=1):
        content = await f.read()
        r = processar_imagem_bytes(f.filename, content)  # deve retornar dict com chaves esperadas
        r["indice_processamento"] = i
        # Garante data_extracao, se o OCR não colocar
        if not r.get("data_extracao"):
            r["data_extracao"] = agora_belem_str()
        resultados.append(r)
        if r.get("protocolo_17_digitos"):
            # conta apenas se tem 17 dígitos válidos após normalização
            if _normaliza_proto_17(r.get("protocolo_17_digitos")):
                encontrados += 1

    # Monta DataFrame do lote
    df_lote = pd.DataFrame([{
        "arquivo": r.get("arquivo", r.get("filename", "")) or "",
        "protocolo_ocr": r.get("protocolo_ocr", ""),
        "protocolo_17_digitos": r.get("protocolo_17_digitos", ""),
        "data_extracao": r.get("data_extracao", agora_belem_str()),
    } for r in resultados])

    # Deduplicação real + gravação
    stats = salvar_excel_sem_duplicar(EXCEL_PATH, df_lote)

    # Lê Excel final e calcula total de protocolos válidos (17 dígitos)
    total_ok_excel = 0
    try:
        df_total = pd.read_excel(EXCEL_PATH, dtype=str, engine="openpyxl")
        if "protocolo_17_digitos" in df_total.columns:
            total_ok_excel = int(
                df_total["protocolo_17_digitos"]
                .astype(str)
                .apply(lambda s: s.isdigit() and len(s) == 17)
                .sum()
            )
    except Exception:
        df_total = pd.DataFrame(columns=COLS)

    resumo = {
        "total_docs": len(resultados),
        "processados": len(resultados),
        "encontrados": encontrados,
        "nao_encontrados": len(resultados) - encontrados,
        "duplicados_no_lote": stats["duplicados_no_lote"],
        "duplicados_no_excel": stats["ja_salvos"],
        "inseridos_no_excel": stats["inseridos"],
        "tempo_total_s": round(time.time() - t0, 3),
    }

    # Compat: mantém chaves usadas no front anterior
    return JSONResponse({
        "resumo": resumo,
        "processados": len(resultados),
        "total_protocolos_no_excel": total_ok_excel,
        "excel_path": "/download/excel",
        "resultados_lote": resultados,
        "duplicados_lote": stats["lista_duplicados_lote"],
        "duplicados": stats["lista_ja_salvos"],   # já existiam no Excel
        "inseridos": stats["lista_inseridos"],    # novos inseridos agora
    })


@app.get("/download/excel")
def download_excel():
    excel = Path(EXCEL_PATH)
    if not excel.exists():
        return JSONResponse({"erro": "Excel ainda não existe"}, status_code=404)
    return FileResponse(
        excel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=excel.name,
    )


# Execução local:  uvicorn app.main_web:app --host 0.0.0.0 --port 8000 --reload
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main_web:app", host="0.0.0.0", port=8000, reload=True)
